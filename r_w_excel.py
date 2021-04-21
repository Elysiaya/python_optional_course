from openpyxl import load_workbook

# 加载excel表选择sheet
wb = load_workbook('人员基本情况.xlsx')
ws = wb.active
# 获取第一列
A = [i.value for i in ws["A"]][1:]

for i, j in enumerate(A):
    c = 'B' + str(i + 2)
    if j is None:
        continue
    else:
        ws[c] = str(j)[:3]

wb.save('人员基本情况.xlsx')
